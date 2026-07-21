from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter



class ExcelManager:
    """
    Handles Excel import and export operations.
    """



    def export_transactions(
        self,
        records,
        file_path
    ):

        workbook = Workbook()


        sheet = workbook.active


        sheet.title = "Transactions"



        headers = [

            "ID",

            "Date",

            "Type",

            "Category",

            "Description",

            "Amount",

            "Created At"

        ]


        sheet.append(
            headers
        )


        for cell in sheet[1]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )



        for record in records:

            sheet.append(
                list(record)
            )



        for row in sheet.iter_rows(
            min_row=2,
            min_col=6,
            max_col=6
        ):

            row[0].number_format = (
                '₹ #,##0.00'
            )



        for column in sheet.columns:

            max_length = 0


            column_letter = get_column_letter(
                column[0].column
            )


            for cell in column:

                if cell.value:

                    max_length = max(
                        max_length,
                        len(
                            str(cell.value)
                        )
                    )


            sheet.column_dimensions[
                column_letter
            ].width = max_length + 3



        workbook.save(
            file_path
        )



    def import_transactions(
        self,
        file_path
    ):
        """
        Reads transactions from Excel.

        Returns list of dictionaries.
        """


        workbook = load_workbook(
            file_path
        )


        sheet = workbook.active


        transactions = []


        for row in sheet.iter_rows(
            min_row=2,
            values_only=True
        ):


            if not row[1]:

                continue



            transaction = {

                "date":
                str(row[1]),


                "type":
                row[2],


                "category":
                row[3],


                "description":
                row[4],


                "amount":
                float(row[5])

            }


            transactions.append(
                transaction
            )


        return transactions